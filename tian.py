# -*- coding: utf-8 -*-
import openpyxl
import xlsxwriter
import xlwt
import xlrd
import xlwings as xw
import random
from openpyxl import load_workbook
import re
import string
import time

def random_letters(n):
    # 定义一个空列表保存随机字母
    letters_list = []
    while len(letters_list) < n:
        a_str = string.ascii_uppercase
        # 字母：string.ascii_letters
        # 大写：string.ascii_uppercase
        # 小写：string.ascii_lowercase
        random_letter = random.choice(a_str)
        if random_letter not in letters_list:
            letters_list.append(random_letter)
        else:
            pass
    # 将列表转换成字符串输出
    letters_list = ''.join(letters_list)
    print(str(letters_list))
    return str(letters_list)


def read_with_dropdown(book_name, sheet_name="Sheet1"):
    # https://blog.csdn.net/weixin_41267342/article/details/86634007。读取下拉列表
    wb_open = openpyxl.load_workbook(book_name)  # 读取excel
    time.sleep(1)

    # 读取sheet表
    ws = wb_open[sheet_name]
    # 读取excel指定单元格数据
    # data = ws["A1":"G5"]
    # 获取内容存在下拉选的框数据

    validations = ws.data_validations.dataValidation
    app = xw.App(visible=False, add_book=False)
    # app = xw.App(spec = 'office')

    wb = app.books.open(book_name)

    sht = wb.sheets['商品属性值']

    # 遍历存在下拉选的单元格
    for validation in validations:
        # 获取下拉框中的所有选择值
        cell = validation.sqref
        result = validation.formula1
        print("单元格位置:" + str(cell) + ",下拉选内容：" + result)
        data_list = sht.range(result).value
        # 进行顺序填充
        tps_front = str(cell)[:str(cell).index(':')]
        tps_later = str(cell)[str(cell).index(':') + 1:]

        # 提取单元格位置中的数字和字母
        tps_front_num = re.sub("\D", "", tps_front)
        tps_later_num = re.sub("\D", "", tps_later)
        tps_front_eng = ''.join(re.findall(r'[A-Za-z]', tps_front))

        for i in range(int(tps_front_num), int(tps_later_num) + 1):
            print('填充数据')
            data = random.choice(data_list)
            print(sht.range(result).value)
            print(data)
            ws[tps_front_eng + str(i)] = data
        # ws[str(cell)] = data

    wb.save()
    wb.close()
    wb_open.save(book_name)


def search_value(filename, keyword):
    wb = load_workbook(filename)
    sheet1 = wb['商品属性']
    product_name = []
    for row in sheet1.iter_rows():
        for cell in row:
            if cell.value is not None:
                info = cell.value.find(keyword)
                if info == 0:
                    # print(cell.value)
                    # print(cell.row, cell.column)
                    value = 'excel入导' + random_letters(3) + str(cell.row + 1)
                    sheet1.cell(row=cell.row + 1, column=cell.column, value=value)
                    # 商品标题和货号保持一致
                    sheet1.cell(row=cell.row + 1, column=cell.column + 1, value=value)
                    # 价格输入1111
                    sheet1.cell(row=cell.row + 1, column=cell.column + 2, value='1111')
                    print(value)
                    product_name.append(value)

    # 把货号复制到sku表里
    sheet2 = wb['sku表']
    print(product_name)
    for row in range(len(product_name)):
        sheet2.cell(row=row + 2, column=2, value=product_name[row])
        # 色号填写
        sheet2.cell(row=row + 2, column=6, value='22')
        # 尺码填写
        sheet2.cell(row=row + 2, column=7, value='xl')
        # 库存填写
        sheet2.cell(row=row + 2, column=8, value='1')
        # 价格填写
        sheet2.cell(row=row + 2, column=9, value='1111')

    wb.save(filename)


if __name__ == '__main__':
    excel_name = '20210517-18_01_29.xlsx'

    read_with_dropdown(excel_name, "商品属性")
    read_with_dropdown(excel_name, "sku表")
    # 详情页文案-推荐尺码表 不填报错
    # read_with_dropdown(excel_name, "详情页文案表")
    try:
        filename = excel_name
        keyword = '货号'
        search_value(filename, keyword)
        print('   Finish')

    except Exception as e:
        print(e)

