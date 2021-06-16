# !/usr/bin/python
# -*- coding: UTF-8 -*-

import pymysql,os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from utils.configures.paths import EXCEL_DATA_DIR


def get_category_prop_info(prop_value_id):
    MYSQL_HOST = 'rm-2zeb07rt531eym8of1o.mysql.rds.aliyuncs.com'
    MYSQL_USERNAME = 'dev'
    MYSQL_PASSWORD = 'lyp82nLF'
    DATABASE_NAME = 'dev_catalog'
    MYSQL_PORT = 3306

    conn = pymysql.connect(host=MYSQL_HOST,
                           user=MYSQL_USERNAME,
                           passwd=MYSQL_PASSWORD,
                           port=MYSQL_PORT,
                           database=DATABASE_NAME)

    cursor = conn.cursor()
    cursor.execute(
        """SELECT category_prop_prop_value.id,category_prop_id,category_prop.category_id FROM category_prop_prop_value
        LEFT JOIN category_prop ON category_prop_prop_value.category_prop_id = category_prop.id
        WHERE prop_value_id=%s""",
        (prop_value_id,))
    data = [(p[0], p[1], p[2]) for p in cursor.fetchall()]
    return data


def parse_xlsx(ws: Workbook):
    data = {}
    for sheet_name in ws.sheetnames:
        sheet = ws[sheet_name]
        columns = [(k.value, k.column) for k in sheet[1] if k.value is not None]
        # 根据prop_value_id归类
        prop_value_map = {}
        for row in sheet:
            if row[0].row == 1 or row[0].value is None:
                continue
            row_info = {}
            for n, idx in columns:
                row_info[n] = row[idx - 1].value

            prop_value_id = row[6].value
            prop_value_map.setdefault(prop_value_id,[]).append(row_info)

        for k, vs in prop_value_map.items():
            # 相同平台归类
            tp_value_map = {}
            for v in vs:
                tp_value_map.setdefault(v['平台名称'], set()).add(v['第三方平台属性值名'])
            # 获取包含类目当前这个属性值的所有类目

            category_info = get_category_prop_info(k)
            for category_prop_value_id, category_prop_id, cid in category_info:
                clear_info = {
                    'category_id': cid,
                    'category_prop_value_id': category_prop_value_id,
                    'category_prop_id': category_prop_id,
                    'jd': ','.join(tp_value_map.get(1003, [])),
                    'tb': ','.join(tp_value_map.get(1000, [])),
                    'tm': ','.join(tp_value_map.get(1001, [])),
                    'vip': ','.join(tp_value_map.get(1005, [])),
                    'pdd': ','.join(tp_value_map.get(1004, []))
                }

                data.setdefault(sheet_name, []).append(clear_info)
    return data

def save_xlsx(filename, info):
    wb = Workbook()
    wss = wb.active
    wss.append(
        ['类目ID', '类目属性ID', '类目属性值ID', 'jd', 'tm', 'tb', 'vip', 'pdd'])
    print(info[0])
    print(returnCategory(info[0]['category_id'])[0],' '*30,filename)

    for i in info:
        wss.append([
            i['category_id'],
            i['category_prop_id'],
            i['category_prop_value_id'],
            i['jd'],
            i['tm'],
            i['tb'],
            i['vip'],
            i['pdd'],
        ])
    wb.save(os.path.join(EXCEL_DATA_DIR,filename + '.xlsx'))

def returnCategory(category_id):
    MYSQL_HOST = 'rm-2zeb07rt531eym8of1o.mysql.rds.aliyuncs.com'
    MYSQL_USERNAME = 'dev'
    MYSQL_PASSWORD = 'lyp82nLF'
    DATABASE_NAME = 'dev_catalog'
    MYSQL_PORT = 3306

    conn = pymysql.connect(host=MYSQL_HOST,
                           user=MYSQL_USERNAME,
                           passwd=MYSQL_PASSWORD,
                           port=MYSQL_PORT,
                           database=DATABASE_NAME)

    cursor = conn.cursor()
    cursor.execute(
        """SELECT path FROM category WHERE id=%s""",
        (category_id,))
    data = cursor.fetchone()
    return data

if __name__ == '__main__':
    props = get_category_prop_info(2133246)
    print(props)